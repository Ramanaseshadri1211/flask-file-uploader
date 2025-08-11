import os
import uuid
from flask import Flask, render_template, request, send_file, jsonify, url_for, Response
from werkzeug.utils import secure_filename
import pdfplumber
import openpyxl
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "outputs"
ALLOWED = {"pdf", "xls", "xlsx"}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200MB

def allowed_file(filename):
    return "." in filename and filename.rsplit(".",1)[1].lower() in ALLOWED

@app.route('/')
def index():
    return render_template("index.html")

@app.route('/pdf-to-excel')
def pdf_to_excel_page():
    return render_template("pdf_to_excel.html")

@app.route('/excel-to-pdf')
def excel_to_pdf_page():
    return render_template("excel_to_pdf.html")

@app.route('/api/pdf-to-excel', methods=['POST'])
def api_pdf_to_excel():
    if 'file' not in request.files:
        return jsonify({"error":"no file"}), 400
    f = request.files['file']
    if f.filename == "" or not allowed_file(f.filename):
        return jsonify({"error":"invalid file"}), 400
    fname = secure_filename(f.filename)
    tmp_path = os.path.join(UPLOAD_FOLDER, f"{uuid.uuid4().hex}_{fname}")
    f.save(tmp_path)
    out_name = f"{uuid.uuid4().hex}.xlsx"
    out_path = os.path.join(OUTPUT_FOLDER, out_name)
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted"
        with pdfplumber.open(tmp_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        for row in table:
                            safe_row = [cell if cell is not None else "" for cell in row]
                            ws.append(safe_row)
                else:
                    text = page.extract_text()
                    if text:
                        for line in text.splitlines():
                            ws.append([line])
        wb.save(out_path)
    except Exception as e:
        try:
            os.remove(tmp_path)
        except:
            pass
        return jsonify({"error": str(e)}), 500
    try:
        os.remove(tmp_path)
    except:
        pass
    return send_file(out_path, as_attachment=True, download_name="converted.xlsx")

@app.route('/api/excel-to-pdf', methods=['POST'])
def api_excel_to_pdf():
    if 'file' not in request.files:
        return jsonify({"error":"no file"}), 400
    f = request.files['file']
    if f.filename == "" or not allowed_file(f.filename):
        return jsonify({"error":"invalid file"}), 400
    fname = secure_filename(f.filename)
    tmp_path = os.path.join(UPLOAD_FOLDER, f"{uuid.uuid4().hex}_{fname}")
    f.save(tmp_path)
    out_name = f"{uuid.uuid4().hex}.pdf"
    out_path = os.path.join(OUTPUT_FOLDER, out_name)
    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        c = canvas.Canvas(out_path, pagesize=A4)
        width, height = A4
        margin = 40
        y = height - margin
        for sheet in wb.worksheets:
            c.setFont("Helvetica-Bold", 12)
            c.drawString(margin, y, f"Sheet: {sheet.title}")
            y -= 18
            c.setFont("Helvetica", 9)
            for r in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in r])
                if y < margin + 40:
                    c.showPage()
                    y = height - margin
                c.drawString(margin, y, (row_text)[:1200])
                y -= 12
            c.showPage()
            y = height - margin
        c.save()
    except Exception as e:
        try:
            os.remove(tmp_path)
        except:
            pass
        return jsonify({"error": str(e)}), 500
    try:
        os.remove(tmp_path)
    except:
        pass
    return send_file(out_path, as_attachment=True, download_name="converted.pdf")

# Serve sitemap and robots for SEO
@app.route('/robots.txt')
def robots():
    lines = [
        "User-agent: *",
        "Allow: /",
        f"Sitemap: {request.url_root.rstrip('/')}/sitemap.xml"
    ]
    return Response("\n".join(lines), mimetype="text/plain")

@app.route('/sitemap.xml')
def sitemap():
    pages = [
        ('/', '1.0'),
        ('/pdf-to-excel', '0.9'),
        ('/excel-to-pdf', '0.9'),
    ]
    xml_parts = ['<?xml version="1.0" encoding="UTF-8"?>',
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
    for p, pr in pages:
        full = request.url_root.rstrip('/') + p
        xml_parts.append(f"<url><loc>{full}</loc><priority>{pr}</priority></url>")
    xml_parts.append('</urlset>')
    return Response("\n".join(xml_parts), mimetype='application/xml')


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
